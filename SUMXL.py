import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.cell import Cell
#Copy range of cells as a nested list
#Takes: start cell, end cell, and sheet you want to copy from.
def copyRange(startCol, startRow, endCol, endRow, sheet):
    rangeSelected = []
    #Loops through selected Rows
    for i in range(startRow,endRow + 1,1):
        #Appends the row to a RowSelected list
        rowSelected = []
        for j in range(startCol,endCol+1,1):
            rowSelected.append(sheet.cell(row = i, column = j).value)
        #Adds the RowSelected List and nests inside the rangeSelected
        rangeSelected.append(rowSelected)
    return rangeSelected

#paste
#Paste range
#Paste data from copyRange into template sheet
def pasteRange(startCol, startRow, endCol, endRow, sheetReceiving,copiedData):
    countRow = 0
    for i in range(startRow,endRow+1,1):
        countCol = 0
        for j in range(startCol,endCol+1,1):
            
            sheetReceiving.cell(row = i, column = j).value = copiedData[countRow][countCol]
            countCol += 1
        countRow += 1

#sum function
def SUMXL(path1,path2,pathsum):
    #new file for sum
    filepath=pathsum
    summ=openpyxl.Workbook()
    summ.save(filepath)
    filesum=openpyxl.load_workbook(pathsum)
    filesum.create_sheet("DASHBOARD")
    filesum.save(pathsum)
    #existing reports
    file1=openpyxl.load_workbook(path1, data_only=True)
    file2=openpyxl.load_workbook(path2, data_only=True)
    #working on dashboard sheet to retrieve data
    sheet1=file1.get_sheet_by_name('DASHBOARD')
    sheet2=file2.get_sheet_by_name('DASHBOARD')
    

    
   

    #write to test
    sheet3=filesum.get_sheet_by_name('DASHBOARD')
    #copypaste
    copydata=copyRange(6,1,9,44,sheet2)
    pasteRange(6,1,9,44,sheet3,copydata)

    #sum
    extraction=sheet1['I6'].value+sheet2['I6'].value
    epierrage=sheet1['I8'].value+sheet2['I8'].value
    volume_phos_extr=sheet1['I10'].value+sheet2['I10'].value
    volume_sterile=sheet1['I12'].value+sheet2['I12'].value
    surface_exploitee=sheet1['I14'].value+sheet2['I14'].value
    ratiotm3=sheet1['I16'].value+sheet2['I16'].value
    ratiom3m3=sheet1['I18'].value+sheet2['I18'].value
    tauxrecup=(sheet1['I20'].value+sheet2['I20'].value)/2
    tauxsaliss=(sheet1['I22'].value+sheet2['I22'].value)/2
    avancephosph=sheet1['I24'].value+sheet2['I24'].value
    avancesautage=sheet1['I26'].value+sheet2['I26'].value
    volumesaut=sheet1['I28'].value+sheet2['I28'].value
    qteexplo=sheet1['I30'].value+sheet2['I30'].value
    dosage=sheet1['I32'].value+sheet2['I32'].value
    tovex=sheet1['I34'].value+sheet2['I34'].value
    dmi=sheet1['I36'].value+sheet2['I36'].value
    vkuocp=sheet1['I38'].value+sheet2['I38'].value
    vkust=sheet1['I40'].value+sheet2['I40'].value
    
    #total
    sheet3['C4'].value='TOTAL'
    sheet3['I6'].value=extraction
    sheet3['I8'].value=epierrage
    sheet3['I10'].value=volume_phos_extr
    sheet3['I12'].value=volume_sterile
    sheet3['I14']=surface_exploitee
    sheet3['I16']=ratiotm3
    sheet3['I18']=ratiom3m3
    sheet3['I20']=tauxrecup
    sheet3['I22']=tauxsaliss
    sheet3['I24']=avancephosph
    sheet3['I26']=avancesautage
    sheet3['I28']=volumesaut
    sheet3['I30']=qteexplo
    sheet3['I32']=dosage
    sheet3['I34']=tovex
    sheet3['I36']=dmi
    sheet3['I38']=vkuocp
    sheet3['I40']=vkust


    #style
    greyFill = PatternFill(start_color='F1F1F1',
                   end_color='F1F1F1',
                   fill_type='solid')
    for i in range(5,42,1):
        for j in range(8,10,1):
            sheet3.cell(row = i, column = j).fill=greyFill

    greenFill = PatternFill(start_color='76B72F',
                   end_color='76B72F',
                   fill_type='solid')

    sheet3['C4'].fill=greenFill
    sheet3['H4'].fill=greenFill
    sheet3['I4'].fill=greenFill



    
    filesum.save(pathsum)
    print extraction

